import openpyxl

def work(name_sheet, dictionary):
    workbook = openpyxl.load_workbook('rjd.xlsx')
    sheet = workbook[name_sheet]
    last_row = sheet.max_row
    for i in dictionary.keys():
        sheet.cell(row=last_row + 1, column=i, value=dictionary[i][1])
    workbook.save('rjd.xlsx')